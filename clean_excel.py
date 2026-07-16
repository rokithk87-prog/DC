"""
clean_excel.py — Core cleaning logic for Excel Cleaner Streamlit app.
"""

import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants & Mappings ──────────────────────────────────────────────────────

WORD_TO_NUM = {
    "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10
}

COUNTRY_MAP = {
    "usa": "USA", "us": "USA", 
    "uk": "UK", "gb": "UK", 
    "uae": "UAE",
    "canada": "Canada", "ca": "Canada",
    "germany": "Germany", "de": "Germany",
    "france": "France", "fr": "France",
    "australia": "Australia", "au": "Australia",
    "india": "India", "in": "India"
}

MISSING_PLACEHOLDERS = {"", "n/a", "null", "none", "-", "nan", "nil", "#n/a"}

def is_missing(val):
    if pd.isna(val):
        return True
    if isinstance(val, str) and val.strip().lower() in MISSING_PLACEHOLDERS:
        return True
    return False

def get_id_gen():
    i = 0
    def gen():
        nonlocal i
        i += 1
        return f"MISSING-{i:03d}"
    return gen

# ── Main cleaning function ────────────────────────────────────────────────────

def clean_dataframe(df: pd.DataFrame):
    """
    Clean *df* in-place (caller should pass a copy).

    Returns
    -------
    df_clean    : pd.DataFrame
    report      : dict
    outlier_mask: pd.DataFrame (bool, same shape as df_clean)
    """
    report = {
        "original_shape": df.shape,
        "final_shape": df.shape,
        "duplicates_removed": 0,
        "empty_rows_removed": 0,
        "empty_cols_removed": 0,
        "columns_renamed": {},
        "date_cols_detected": [],
        "numeric_cols_converted": [],
        "outliers_flagged": {},
        "fixed_counts": {}
    }
    
    df.columns = [str(c).strip() for c in df.columns]

    before_rows = len(df)
    df.dropna(how="all", inplace=True)
    report["blank_rows_removed"] = before_rows - len(df)

    before_dupes = len(df)
    df.drop_duplicates(inplace=True)
    report["duplicates_removed"] = before_dupes - len(df)

    fixed_counts = {
        'Order ID': 0, 'Customer Name': 0, 'Email': 0, 'Phone': 0, 'Country': 0,
        'Order Date': 0, 'Product': 0, 'Quantity': 0, 'Unit Price': 0, 'Total': 0,
        'Status': 0, 'Sales Rep': 0
    }

    missing_id_gen = get_id_gen()
    cleaned_rows = []

    for _, row in df.iterrows():
        issues = []
        fixed = set()

        # ── Order ID ──
        oid = row.get('Order ID', np.nan)
        if is_missing(oid):
            issues.append("Generated missing Order ID")
            fixed.add('Order ID')
            oid = missing_id_gen()
        else:
            cleaned_oid = re.sub(r'\s+', '', str(oid).upper())
            if cleaned_oid != str(oid).strip():
                fixed.add('Order ID')
            oid = cleaned_oid

        # ── Customer Name ──
        cname = row.get('Customer Name', np.nan)
        if is_missing(cname):
            issues.append("Missing customer name")
            fixed.add('Customer Name')
            cname = "Unknown Customer"
        else:
            cleaned_cname = str(cname).strip().title()
            if cleaned_cname != str(cname).strip():
                fixed.add('Customer Name')
            cname = cleaned_cname

        # ── Email ──
        email = row.get('Email', np.nan)
        if is_missing(email):
            issues.append("Missing email")
            fixed.add('Email')
            email = "missing@unknown.com"
        else:
            cleaned_email = str(email).strip().lower()
            valid = True
            if "@@" in cleaned_email:
                valid = False
            parts = cleaned_email.split('@')
            if len(parts) != 2 or not parts[0] or not parts[1]:
                valid = False
            if '.' not in parts[-1]:
                valid = False

            if not valid:
                issues.append("Invalid email")
                fixed.add('Email')
                email = "invalid@unknown.com"
            elif cleaned_email != str(email).strip().lower():
                fixed.add('Email')

        # ── Phone ──
        phone = row.get('Phone', np.nan)
        if is_missing(phone):
            issues.append("Missing phone")
            fixed.add('Phone')
            phone = "N/A"
        else:
            digits = re.sub(r'\D', '', str(phone))
            if digits == "0000000000" or str(phone).strip() == "000-000-0000":
                issues.append("Missing phone")
                fixed.add('Phone')
                phone = "N/A"
            elif len(digits) == 10:
                cleaned_phone = f"+1-{digits[:3]}-{digits[3:6]}-{digits[6:]}"
                if cleaned_phone != str(phone).strip():
                    fixed.add('Phone')
                phone = cleaned_phone
            elif len(digits) == 11 and digits.startswith('1'):
                cleaned_phone = f"+1-{digits[1:4]}-{digits[4:7]}-{digits[7:]}"
                if cleaned_phone != str(phone).strip():
                    fixed.add('Phone')
                phone = cleaned_phone
            else:
                issues.append("Invalid phone")
                fixed.add('Phone')
                phone = "N/A"

        # ── Country ──
        country = row.get('Country', np.nan)
        if is_missing(country) or str(country).strip().lower() == 'xx':
            issues.append("Missing country")
            fixed.add('Country')
            country = "Unknown"
        else:
            c = str(country).strip().lower()
            if c in COUNTRY_MAP:
                cleaned_country = COUNTRY_MAP[c]
            else:
                cleaned_country = str(country).strip().title()
            if cleaned_country != str(country).strip():
                fixed.add('Country')
            country = cleaned_country

        # ── Order Date ──
        odate = row.get('Order Date', np.nan)
        if is_missing(odate):
            issues.append("Missing date")
            fixed.add('Order Date')
            odate = "1900-01-01"
        else:
            odate_str = str(odate).strip()
            parsed_date = pd.NaT

            match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', odate_str)
            if match and int(match.group(1)) > 12:
                parsed_date = pd.to_datetime(odate_str, dayfirst=True, errors='coerce')
            else:
                parsed_date = pd.to_datetime(odate_str, errors='coerce')
                if pd.isna(parsed_date):
                    parsed_date = pd.to_datetime(odate_str, dayfirst=True, errors='coerce')

            if pd.isna(parsed_date):
                issues.append("Invalid date")
                fixed.add('Order Date')
                odate = "1900-01-01"
            else:
                cleaned_odate = parsed_date.strftime("%Y-%m-%d")
                if cleaned_odate != odate_str:
                    fixed.add('Order Date')
                odate = cleaned_odate

        # ── Product ──
        prod = row.get('Product', np.nan)
        if is_missing(prod):
            issues.append("Missing product")
            fixed.add('Product')
            prod = "Unknown Product"
        else:
            cleaned_prod = str(prod).strip().title()
            if cleaned_prod != str(prod).strip():
                fixed.add('Product')
            prod = cleaned_prod

        # ── Quantity ──
        qty = row.get('Quantity', np.nan)
        if is_missing(qty):
            issues.append("Assumed quantity=1")
            fixed.add('Quantity')
            qty = 1
        else:
            qty_str = str(qty).strip().lower()
            if qty_str in WORD_TO_NUM:
                fixed.add('Quantity')
                qty = WORD_TO_NUM[qty_str]
            else:
                try:
                    val_str = qty_str.replace(',', '')
                    val = float(val_str)
                    if val < 0:
                        fixed.add('Quantity')
                        val = abs(val)
                    if val == 0:
                        issues.append("Assumed quantity=1")
                        fixed.add('Quantity')
                        val = 1
                    if not val.is_integer():
                        fixed.add('Quantity')
                    cleaned_qty = int(val)
                    if qty_str != str(cleaned_qty):
                        fixed.add('Quantity')
                    qty = cleaned_qty
                except ValueError:
                    issues.append("Assumed quantity=1")
                    fixed.add('Quantity')
                    qty = 1

        # ── Unit Price ──
        uprice = row.get('Unit Price', np.nan)
        if is_missing(uprice) or str(uprice).strip().upper() == "TBD":
            issues.append("Assumed unit price=0.0")
            fixed.add('Unit Price')
            uprice = 0.0
        else:
            up_str = str(uprice).strip().lower()
            cleaned_up_str = re.sub(r'[\$£€]', '', up_str)
            cleaned_up_str = re.sub(r'\busd\b|\baed\b|\bgbp\b|\beur\b', '', cleaned_up_str)
            cleaned_up_str = cleaned_up_str.replace(',', '').strip()

            try:
                val = float(cleaned_up_str)
                if val < 0:
                    fixed.add('Unit Price')
                    val = abs(val)

                cleaned_up = round(val, 2)
                if cleaned_up == 0.0:
                    issues.append("Assumed unit price=0.0")
                    fixed.add('Unit Price')

                if str(uprice).strip().lower() != f"{cleaned_up:.2f}":
                    fixed.add('Unit Price')
                uprice = cleaned_up
            except ValueError:
                issues.append("Assumed unit price=0.0")
                fixed.add('Unit Price')
                uprice = 0.0

        # ── Total ──
        total_orig = row.get('Total', np.nan)
        total = round(qty * uprice, 2)

        if is_missing(total_orig) or str(total_orig).strip().upper() == "TBD":
            fixed.add('Total')
        else:
            tot_str = str(total_orig).strip().lower()
            tot_str_clean = re.sub(r'[\$£€]', '', tot_str)
            tot_str_clean = re.sub(r'\busd\b|\baed\b|\bgbp\b|\beur\b', '', tot_str_clean)
            tot_str_clean = tot_str_clean.replace(',', '').strip()

            try:
                val = float(tot_str_clean)
                val_rounded = round(abs(val), 2)
                if val_rounded != total or val < 0:
                    fixed.add('Total')
                if tot_str != f"{total:.2f}":
                    fixed.add('Total')
            except ValueError:
                fixed.add('Total')

        # ── Status ──
        status = row.get('Status', np.nan)
        valid_statuses = {"Completed", "Pending", "Shipped", "Refunded", "Cancelled"}
        st_raw = str(status).strip() if not is_missing(status) else ""

        if is_missing(status) or st_raw in {"???", "NULL"}:
            issues.append("Unknown status")
            fixed.add('Status')
            status = "Unknown"
        else:
            st = st_raw.title()
            if st in valid_statuses:
                if st != st_raw:
                    fixed.add('Status')
                status = st
            else:
                issues.append("Unknown status")
                fixed.add('Status')
                status = "Unknown"

        # ── Sales Rep ──
        srep = row.get('Sales Rep', np.nan)
        if is_missing(srep):
            issues.append("Unassigned sales rep")
            fixed.add('Sales Rep')
            srep = "Unassigned"
        else:
            cleaned_srep = str(srep).strip().title()
            if cleaned_srep != str(srep).strip():
                fixed.add('Sales Rep')
            srep = cleaned_srep

        for col in fixed:
            if col in fixed_counts:
                fixed_counts[col] += 1

        cleaned_rows.append([
            oid, cname, email, phone, country, odate, prod, qty, uprice, total, status, srep, ", ".join(issues)
        ])

    cleaned_df = pd.DataFrame(cleaned_rows, columns=[
        'Order ID', 'Customer Name', 'Email', 'Phone', 'Country', 'Order Date',
        'Product', 'Quantity', 'Unit Price', 'Total', 'Status', 'Sales Rep', 'Data Quality Issues'
    ])
    
    cleaned_df['Quantity'] = pd.to_numeric(cleaned_df['Quantity'], errors='coerce').astype(int)
    cleaned_df['Unit Price'] = pd.to_numeric(cleaned_df['Unit Price'], errors='coerce')
    cleaned_df['Total'] = pd.to_numeric(cleaned_df['Total'], errors='coerce')

    report["fixed_counts"] = fixed_counts
    report["final_shape"] = cleaned_df.shape
    report["outliers_flagged"] = {}  # Not used in strict rules
    
    # Return an empty outlier_mask to maintain compatibility with app.py and write_cleaned_excel
    outlier_mask = pd.DataFrame(False, index=cleaned_df.index, columns=cleaned_df.columns)

    return cleaned_df, report, outlier_mask


# ── Excel writer ──────────────────────────────────────────────────────────────

def _header_style(ws, n_cols: int):
    fill = PatternFill("solid", fgColor="1E293B")
    font = Font(bold=True, color="F1F5F9", name="Calibri", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="334155")
    border = Border(bottom=thin)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

# FIX: Removed `pd.DataFrame | None` type hint for Python 3.8/3.9 compatibility
def _df_to_sheet(ws, df: pd.DataFrame, *, highlight_mask=None):
    """Write df to ws with optional cell highlighting."""
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=j, value=str(col))
    _header_style(ws, len(df.columns))

    orange_fill = PatternFill("solid", fgColor="7C2D12")
    alt_fill    = PatternFill("solid", fgColor="0F172A")
    alt_fill2   = PatternFill("solid", fgColor="0F1D2E")
    default_font = Font(name="Calibri", size=10, color="CBD5E1")

    col_index = {col: j for j, col in enumerate(df.columns, 1)}

    for i, (_, row) in enumerate(df.iterrows(), 2):
        row_fill = alt_fill if i % 2 == 0 else alt_fill2
        for col in df.columns:
            j = col_index[col]
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = default_font
            cell.fill = row_fill
            # Safe check for highlight_mask
            if highlight_mask is not None and col in highlight_mask.columns and highlight_mask.at[i - 2, col]:
                cell.fill = orange_fill

    ws.freeze_panes = "A2"
    _auto_width(ws)

def write_cleaned_excel(
    df_clean: pd.DataFrame,
    report: dict,
    outlier_mask: pd.DataFrame,
    path: str,
):
    """Write a 3-sheet Excel workbook: Cleaned Data, Outliers, Cleaning Report."""
    wb = Workbook()

    # ── Sheet 1: Cleaned Data ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cleaned Data"
    ws1.sheet_view.showGridLines = False
    _df_to_sheet(ws1, df_clean, highlight_mask=outlier_mask)

    # ── Sheet 2: Outliers ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Outliers")
    ws2.sheet_view.showGridLines = False
    outlier_rows = df_clean[outlier_mask.any(axis=1)].copy()
    if outlier_rows.empty:
        ws2.cell(row=1, column=1, value="No outliers detected.")
    else:
        _df_to_sheet(ws2, outlier_rows)

    # ── Sheet 3: Cleaning Report ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Cleaning Report")
    ws3.sheet_view.showGridLines = False

    fill_hdr = PatternFill("solid", fgColor="1E293B")
    fill_row = PatternFill("solid", fgColor="0F172A")
    font_hdr = Font(bold=True, color="7DD3FC", name="Calibri", size=10)
    font_val = Font(color="CBD5E1", name="Calibri", size=10)
    font_key = Font(bold=True, color="94A3B8", name="Calibri", size=10)

    def _write(r, c, val, bold=False, header=False):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.fill = fill_hdr if header else fill_row
        cell.font = font_hdr if header else (font_key if bold else font_val)
        cell.alignment = Alignment(vertical="center")
        return cell

    rows = [
        ("Metric", "Value"),
        ("Original rows", report["original_shape"][0]),
        ("Original columns", report["original_shape"][1]),
        ("Final rows", report["final_shape"][0]),
        ("Final columns", report["final_shape"][1]),
        ("Duplicates removed", report["duplicates_removed"]),
        ("Empty rows removed", report["empty_rows_removed"]),
        ("Empty columns removed", report["empty_cols_removed"]),
        ("Date columns detected", ", ".join(report.get("date_cols_detected", [])) or "None"),
        ("Numeric cols converted", ", ".join(report.get("numeric_cols_converted", [])) or "None"),
        ("Columns renamed", str(report.get("columns_renamed", {})) if report.get("columns_renamed") else "None"),
        ("Outliers flagged", str(report.get("outliers_flagged", {})) if report.get("outliers_flagged") else "None"),
        ("", ""),
        ("Cells Fixed Per Column", ""),
    ]

    for col, count in report.get("fixed_counts", {}).items():
        rows.append((col, count))

    for r_idx, (key, val) in enumerate(rows, 1):
        header = r_idx == 1
        _write(r_idx, 1, key, bold=not header, header=header)
        _write(r_idx, 2, val, header=header)

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 60
    ws3.freeze_panes = "A2"

    wb.save(path)

# Aliases — app.py uses both names across versions
write_output = write_cleaned_excel
